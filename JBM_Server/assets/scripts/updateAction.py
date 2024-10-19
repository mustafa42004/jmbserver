from openpyxl import load_workbook
import sys

def update_xlsx(file_path, agreement_number, action_status, action_time):
    # Load the existing workbook
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Determine column indices for ACTION, HOLD, RELEASE, and IN_YARD
    header_row = sheet[1]
    action_col_idx = None
    hold_col_idx = None
    release_col_idx = None
    in_yard_col_idx = None

    # Find column indices or create columns if not present
    max_col_idx = len(header_row)  # Get the current max number of columns
    for col_idx, cell in enumerate(header_row, 1):
        if cell.value == 'ACTION':
            action_col_idx = col_idx
        elif cell.value == 'HOLD':
            hold_col_idx = col_idx
        elif cell.value == 'RELEASE':
            release_col_idx = col_idx
        elif cell.value == 'IN_YARD':
            in_yard_col_idx = col_idx

    # Create columns if they don't exist
    if action_col_idx is None:
        action_col_idx = max_col_idx + 1
        sheet.cell(row=1, column=action_col_idx, value='ACTION')
        max_col_idx += 1
    if hold_col_idx is None:
        hold_col_idx = max_col_idx + 1
        sheet.cell(row=1, column=hold_col_idx, value='HOLD')
        max_col_idx += 1
    if release_col_idx is None:
        release_col_idx = max_col_idx + 1
        sheet.cell(row=1, column=release_col_idx, value='RELEASE')
        max_col_idx += 1
    if in_yard_col_idx is None:
        in_yard_col_idx = max_col_idx + 1
        sheet.cell(row=1, column=in_yard_col_idx, value='IN_YARD')
        max_col_idx += 1

    # Update the ACTION, HOLD, RELEASE, or IN_YARD column for the specific agreement_number
    found_agreement = False  # To track if we found the agreement number
    for row in sheet.iter_rows(min_row=2):
        cell_value = row[0].value  # Assuming the agreement number is in the first column
        if cell_value == agreement_number:
            found_agreement = True
            # Update ACTION field
            row[action_col_idx - 1].value = action_status

            # Update the time based on action_status (case insensitive)
            if action_status.lower() == 'hold':
                row[hold_col_idx - 1].value = action_time
            elif action_status.lower() == 'release':
                row[release_col_idx - 1].value = action_time
            elif action_status.lower() == 'in yard':
                row[in_yard_col_idx - 1].value = action_time
            break

    if not found_agreement:
        print(f"Agreement number '{agreement_number}' not found.")

    # Save the workbook with changes
    workbook.save(file_path)
    print(f"File '{file_path}' updated successfully.")

if __name__ == "__main__":
    # Validate arguments
    if len(sys.argv) != 5:
        print("Usage: python updateAction.py <file_path> <agreement_number> <action_status> <action_time>")
        sys.exit(1)

    file_path = sys.argv[1]          # Path to the XLSX file
    agreement_number = sys.argv[2]   # Agreement number to find
    action_status = sys.argv[3]      # New action status to set
    action_time = sys.argv[4]        # Time to be recorded

    update_xlsx(file_path, agreement_number, action_status, action_time)
