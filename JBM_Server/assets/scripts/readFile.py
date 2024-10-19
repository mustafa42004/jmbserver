import pandas as pd
from openpyxl import load_workbook
import sys
import os

# Define the path to the Excel file
# path = r"D:\Inern-task\JMB-Repo\JBM Enterprises\public\assets\samplefiles\Tent Repo List - Jan 24.xlsx"
path = r"D:\Inern-task\JMB-Repo\JBM Enterprises\public\assets\samplefiles\Updated_Tent_Repo_List.xlsx"

if os.path.exists(path):
    print("File exists")
else:
    print("File does not exist")


# def update_xlsx(file_path, filename, bank_name, new_file_name=None):
#     try:
#         if not os.path.isfile(file_path):
#             print(f"File does not exist: {file_path}")
#             return

#         print(f"Updating file: {file_path} with filename: {filename} and bank: {bank_name}")

#         # Load the existing workbook
#         workbook = load_workbook(file_path)
#         sheet = workbook.active

#         # Check if headers exist, and if not, add them
#         headers = [cell.value for cell in sheet[1]]
#         print(f"Headers found: {headers}")

#         # Adding missing headers if they don't exist
#         if 'FILENAME' not in headers:
#             headers.append('FILENAME')
#             sheet.cell(row=1, column=len(headers), value='FILENAME')
        
#         if 'BANK' not in headers:
#             headers.append('BANK')
#             sheet.cell(row=1, column=len(headers), value='BANK')

#         if 'HOLD' not in headers:
#             headers.append('HOLD')
#             sheet.cell(row=1, column=len(headers), value='HOLD')
        
#         if 'RELEASE' not in headers:
#             headers.append('RELEASE')
#             sheet.cell(row=1, column=len(headers), value='RELEASE')

#         if 'IN_YARD' not in headers:
#             headers.append('IN_YARD')
#             sheet.cell(row=1, column=len(headers), value='IN_YARD')

#         if 'ACTION' not in headers:
#             headers.append('ACTION')
#             sheet.cell(row=1, column=len(headers), value='ACTION')

#         # Get column indices after adding the headers
#         filename_col = headers.index('FILENAME') + 1
#         bank_col = headers.index('BANK') + 1
#         hold_col = headers.index('HOLD') + 1
#         release_col = headers.index('RELEASE') + 1
#         in_yard_col = headers.index('IN_YARD') + 1
#         action_col = headers.index('ACTION') + 1

#         # Iterate over rows and add/update 'FILENAME', 'BANK', 'HOLD', 'RELEASE', 'IN_YARD', and 'ACTION'
#         for row in sheet.iter_rows(min_row=2):
#             row[filename_col-1].value = filename   # Update FILENAME
#             row[bank_col-1].value = bank_name      # Update BANK
#             row[hold_col-1].value = 'empty'        # Update HOLD to 'empty'
#             row[release_col-1].value = 'empty'     # Update RELEASE to 'empty'
#             row[in_yard_col-1].value = 'empty'     # Update IN_YARD to 'empty'
#             row[action_col-1].value = 'empty'      # Update ACTION to 'empty'

#         # Define the new file path if a new file name is provided, else save to the original path
#         if new_file_name:
#             new_file_path = os.path.join(os.path.dirname(file_path), new_file_name)
#             print(f"Saving updated file as: {new_file_path}")
#         else:
#             new_file_path = file_path  # If no new file name, overwrite the existing file

#         # Save the workbook with changes
#         try:
#             workbook.save(new_file_path)
#             print(f"File updated successfully: {new_file_path}")
#         except Exception as e:
#             print(f"Error while saving the file: {e}")
    
#     except Exception as e:
#         print(f"Error updating file: {e}")

# if __name__ == "__main__":
#     if len(sys.argv) < 4:
#         print("Usage: python update_xlsx.py <file_path> <filename> <bank_name> [new_file_name]")
#         sys.exit(1)

#     filename = "hhdfc"
#     bank_name = "Kotak"
    
#     # Optional new file name to save as
#     new_file_name = "Updated_Tent_Repo_List.xlsx"  # You can change this to any desired new name
    
#     # Call the function to update and save the file
#     update_xlsx(path, filename, bank_name, new_file_name)




def update_xlsx(file_path, agreement_number, action_status, action_time):
    # Load the existing workbook
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Determine column indices for ACTION, HOLD, RELEASE, and IN_YARD
    header_row = sheet[1]
    action_col_idx = None
    hold_col_idx = None
    release_col_idx = None
    in_yard_col_idx = None

    # Find column indices or create columns if not present
    max_col_idx = len(header_row)  # Get the current max number of columns
    for col_idx, cell in enumerate(header_row, 1):
        if cell.value == 'ACTION':
            action_col_idx = col_idx
        elif cell.value == 'HOLD':
            hold_col_idx = col_idx
        elif cell.value == 'RELEASE':
            release_col_idx = col_idx
        elif cell.value == 'IN_YARD':
            in_yard_col_idx = col_idx

    # Create columns if they don't exist
    if action_col_idx is None:
        action_col_idx = max_col_idx + 1
        sheet.cell(row=1, column=action_col_idx, value='ACTION')
        max_col_idx += 1
    if hold_col_idx is None:
        hold_col_idx = max_col_idx + 1
        sheet.cell(row=1, column=hold_col_idx, value='HOLD')
        max_col_idx += 1
    if release_col_idx is None:
        release_col_idx = max_col_idx + 1
        sheet.cell(row=1, column=release_col_idx, value='RELEASE')
        max_col_idx += 1
    if in_yard_col_idx is None:
        in_yard_col_idx = max_col_idx + 1
        sheet.cell(row=1, column=in_yard_col_idx, value='IN_YARD')
        max_col_idx += 1

    # Update the ACTION, HOLD, RELEASE, or IN_YARD column for the specific agreement_number
    found_agreement = False  # To track if we found the agreement number
    for row in sheet.iter_rows(min_row=2):
        cell_value = row[0].value  # Assuming the agreement number is in the first column
        if cell_value == agreement_number:
            found_agreement = True
            # Update ACTION field
            row[action_col_idx - 1].value = action_status

            # Update the time based on action_status (case insensitive)
            if action_status.lower() == 'hold':
                row[hold_col_idx - 1].value = action_time
            elif action_status.lower() == 'release':
                row[release_col_idx - 1].value = action_time
            elif action_status.lower() == 'in yard':
                row[in_yard_col_idx - 1].value = action_time
            break

    if not found_agreement:
        print(f"Agreement number '{agreement_number}' not found.")

    # Save the workbook with changes
    workbook.save(file_path)
    print(f"File '{file_path}' updated successfully.")

# if __name__ == "__main__":
#     # Validate arguments
#     if len(sys.argv) != 5:
#         print("Usage: python updateAction.py <file_path> <agreement_number> <action_status> <action_time>")
#         sys.exit(1)

file_path = path           # Path to the XLSX file
agreement_number = 'X0CESNG00003776789'  # Agreement number to find
action_status = 'hold'      # New action status to set
action_time =  'monday'       # Time to be recorded

update_xlsx(file_path, agreement_number, action_status, action_time)