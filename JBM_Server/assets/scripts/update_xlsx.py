from openpyxl import load_workbook
import sys
import os

def update_xlsx(file_path, filename, bank_name):
    try:
        if not os.path.isfile(file_path):
            print(f"File does not exist: {file_path}")
            return

        print(f"Updating file: {file_path} with filename: {filename} and bank: {bank_name}")

        # Load the existing workbook
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Check if headers exist, and if not, add them
        headers = [cell.value for cell in sheet[1]]
        print(f"Headers found: {headers}")

        # Adding missing headers if they don't exist
        if 'FILENAME' not in headers:
            headers.append('FILENAME')
            sheet.cell(row=1, column=len(headers), value='FILENAME')
        
        if 'BANK' not in headers:
            headers.append('BANK')
            sheet.cell(row=1, column=len(headers), value='BANK')

        if 'HOLD' not in headers:
            headers.append('HOLD')
            sheet.cell(row=1, column=len(headers), value='HOLD')
        
        if 'RELEASE' not in headers:
            headers.append('RELEASE')
            sheet.cell(row=1, column=len(headers), value='RELEASE')

        if 'IN_YARD' not in headers:
            headers.append('IN_YARD')
            sheet.cell(row=1, column=len(headers), value='IN_YARD')

        if 'ACTION' not in headers:
            headers.append('ACTION')
            sheet.cell(row=1, column=len(headers), value='ACTION')

        # Get column indices after adding the headers
        filename_col = headers.index('FILENAME') + 1
        bank_col = headers.index('BANK') + 1
        hold_col = headers.index('HOLD') + 1
        release_col = headers.index('RELEASE') + 1
        in_yard_col = headers.index('IN_YARD') + 1
        action_col = headers.index('ACTION') + 1

        # Iterate over rows and add/update 'FILENAME', 'BANK', 'HOLD', 'RELEASE', 'IN_YARD', and 'ACTION'
        for row in sheet.iter_rows(min_row=2):
            row[filename_col-1].value = filename   # Update FILENAME
            row[bank_col-1].value = bank_name      # Update BANK
            row[hold_col-1].value = 'empty'        # Update HOLD to 'empty'
            row[release_col-1].value = 'empty'     # Update RELEASE to 'empty'
            row[in_yard_col-1].value = 'empty'     # Update IN_YARD to 'empty'
            row[action_col-1].value = 'empty'      # Update ACTION to 'empty'

        # Save the workbook with changes
        try:
            workbook.save(file_path)
            print(f"File updated successfully: {file_path}")
        except Exception as e:
            print(f"Error while saving the file: {e}")
    
    except Exception as e:
        print(f"Error updating file: {e}")

if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Usage: python update_xlsx.py <file_path> <filename> <bank_name>")
        sys.exit(1)

    file_path = sys.argv[1]  # Path to the XLSX file
    filename = sys.argv[2]   # FILENAME to be added
    bank_name = sys.argv[3]  # Bank name to be added
    update_xlsx(file_path, filename, bank_name)
